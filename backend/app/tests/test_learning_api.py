from collections.abc import Iterator
from datetime import UTC, datetime, timedelta
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.config import settings
from app.core.database import Base, get_db
from app.core.security import create_candidate_token, create_session_token
from app.main import create_app
from app.models import Candidate, ExamAttempt, ExamCandidateScope, LearningVideo
from app.services.operational_lock_service import acquire_backup_write_freeze
from app.tests.conftest import (
    create_candidate,
    create_exam,
    create_question_with_options,
)


def _build_client() -> tuple[TestClient, Session]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session_local = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    db = session_local()
    app = create_app()

    def override_get_db() -> Iterator[Session]:
        yield db

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), db


def _admin_headers() -> dict[str, str]:
    return {"X-Admin-Token": create_session_token(settings.admin_username)}


def _candidate_headers(candidate_id: int) -> dict[str, str]:
    return {"X-Candidate-Token": create_candidate_token(candidate_id)}


def _upload_video(
    client: TestClient,
    *,
    title: str = "安全培训",
    filename: str = "training.mp4",
    content_type: str = "video/mp4",
    duration_seconds: int = 100,
) -> dict:
    response = client.post(
        "/api/admin/learning/videos",
        headers=_admin_headers(),
        data={"title": title, "duration_seconds": str(duration_seconds)},
        files={"file": (filename, b"video-bytes", content_type)},
    )
    assert response.status_code == 200
    return response.json()["data"]


def test_admin_upload_learning_video_generates_storage_key() -> None:
    client, db = _build_client()

    video = _upload_video(client)

    assert video["title"] == "安全培训"
    assert video["original_filename"] == "training.mp4"
    assert video["content_type"] == "video/mp4"
    assert video["file_size_bytes"] == len(b"video-bytes")
    assert video["completion_threshold_percent"] == 90
    assert video["status"] == "draft"
    assert video["storage_key"] != "training.mp4"
    assert video["playback_url"].startswith("/media/learning/")
    assert (db.query(LearningVideo).one()).storage_key == video["storage_key"]


def test_admin_upload_rejects_invalid_video_file() -> None:
    client, db = _build_client()

    response = client.post(
        "/api/admin/learning/videos",
        headers=_admin_headers(),
        data={"title": "安全培训", "duration_seconds": "100"},
        files={"file": ("training.txt", b"not-video", "text/plain")},
    )

    assert response.status_code == 400
    assert db.query(LearningVideo).count() == 0


def test_formal_attempt_blocks_video_upload_before_media_or_row_persistence() -> None:
    client, db = _build_client()
    exam = create_exam(db)
    candidate = create_candidate(db)
    now = datetime.now(UTC)
    db.add(
        ExamAttempt(
            exam_id=exam.id,
            candidate_id=candidate.id,
            status="in_progress",
            started_at=now,
            ends_at=now + timedelta(hours=1),
            duration_minutes_snapshot=60,
        )
    )
    db.commit()

    response = client.post(
        "/api/admin/learning/videos",
        headers=_admin_headers(),
        data={"title": "考试中禁止上传", "duration_seconds": "100"},
        files={"file": ("blocked.mp4", b"video-bytes", "video/mp4")},
    )

    assert response.status_code == 409
    assert "正式考试" in response.json()["detail"]
    assert db.query(LearningVideo).count() == 0


def test_backup_freeze_blocks_progress_but_keeps_video_reads_available() -> None:
    client, db = _build_client()
    video = _upload_video(client)
    client.post(
        f"/api/admin/learning/videos/{video['id']}/publish", headers=_admin_headers()
    )
    candidate = create_candidate(db)
    acquire_backup_write_freeze(db, owner="api-backup", ttl_seconds=600)
    db.commit()

    listed = client.get(
        "/api/learning/videos", headers=_candidate_headers(candidate.id)
    )
    progress = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=_candidate_headers(candidate.id),
        json={
            "current_position_seconds": 10,
            "watched_start_seconds": 0,
            "watched_end_seconds": 10,
        },
    )

    assert listed.status_code == 200
    assert progress.status_code == 409
    assert "配对备份" in progress.json()["detail"]


def test_candidate_learning_videos_require_active_candidate_token() -> None:
    client, db = _build_client()
    video = _upload_video(client)
    client.post(
        f"/api/admin/learning/videos/{video['id']}/publish", headers=_admin_headers()
    )
    inactive = Candidate(name="停用人员", status="inactive")
    db.add(inactive)
    db.commit()

    missing_token = client.get("/api/learning/videos")
    inactive_token = client.get(
        "/api/learning/videos", headers=_candidate_headers(inactive.id)
    )

    assert missing_token.status_code == 401
    assert inactive_token.status_code == 404


def test_candidate_sees_only_published_learning_videos() -> None:
    client, db = _build_client()
    candidate = create_candidate(db, name="学习人员", status="active")
    draft = _upload_video(client, title="草稿视频")
    published = _upload_video(client, title="公开视频")
    archived = _upload_video(client, title="归档视频")
    client.post(
        f"/api/admin/learning/videos/{published['id']}/publish",
        headers=_admin_headers(),
    )
    client.post(
        f"/api/admin/learning/videos/{archived['id']}/archive",
        headers=_admin_headers(),
    )

    response = client.get(
        "/api/learning/videos", headers=_candidate_headers(candidate.id)
    )

    assert response.status_code == 200
    titles = [row["title"] for row in response.json()["data"]]
    assert titles == ["公开视频"]
    assert draft["title"] not in titles
    assert archived["title"] not in titles


def test_learning_progress_completion_skips_jumps_and_deduplicates_intervals() -> None:
    client, db = _build_client()
    candidate = create_candidate(db, name="学习人员", status="active")
    video = _upload_video(client, duration_seconds=100)
    client.post(
        f"/api/admin/learning/videos/{video['id']}/publish", headers=_admin_headers()
    )
    headers = _candidate_headers(candidate.id)

    first = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=headers,
        json={
            "current_position_seconds": 30,
            "watched_start_seconds": 0,
            "watched_end_seconds": 30,
        },
    )
    repeated = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=headers,
        json={
            "current_position_seconds": 30,
            "watched_start_seconds": 0,
            "watched_end_seconds": 30,
        },
    )
    jump = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=headers,
        json={
            "current_position_seconds": 95,
            "watched_start_seconds": 95,
            "watched_end_seconds": 95,
        },
    )
    second = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=headers,
        json={
            "current_position_seconds": 60,
            "watched_start_seconds": 30,
            "watched_end_seconds": 60,
        },
    )
    completed = client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=headers,
        json={
            "current_position_seconds": 90,
            "watched_start_seconds": 60,
            "watched_end_seconds": 90,
        },
    )

    assert first.status_code == 200
    assert first.json()["data"]["completion_percent"] == 30
    assert repeated.json()["data"]["completion_percent"] == 30
    assert jump.json()["data"]["completion_percent"] == 30
    assert second.json()["data"]["completion_percent"] == 60
    assert completed.json()["data"]["completion_percent"] == 90
    assert completed.json()["data"]["completed_at"] is not None


def test_admin_learning_report_and_export() -> None:
    client, db = _build_client()
    candidate = create_candidate(
        db,
        name="学习人员",
        employee_no="L001",
        department="安全部",
        status="active",
    )
    video = _upload_video(client, duration_seconds=100)
    client.post(
        f"/api/admin/learning/videos/{video['id']}/publish", headers=_admin_headers()
    )
    client.post(
        f"/api/learning/videos/{video['id']}/progress",
        headers=_candidate_headers(candidate.id),
        json={
            "current_position_seconds": 90,
            "watched_start_seconds": 0,
            "watched_end_seconds": 90,
        },
    )

    report = client.get("/api/admin/learning/reports", headers=_admin_headers())
    export = client.get("/api/admin/learning/reports/export", headers=_admin_headers())

    assert report.status_code == 200
    row = report.json()["data"][0]
    assert row["candidate_name"] == "学习人员"
    assert row["video_title"] == "安全培训"
    assert row["completion_status"] == "in_progress"
    assert row["completion_percent"] == 30
    assert export.status_code == 200
    workbook = load_workbook(BytesIO(export.content))
    assert workbook.active.title == "视频学习"
    assert workbook.active.cell(1, 1).value == "CID · 人员ID"


def test_video_learning_does_not_gate_exam_start_or_submit() -> None:
    client, db = _build_client()
    candidate = create_candidate(db, name="考试人员", status="active")
    exam = create_exam(db, title="独立考试")
    create_question_with_options(db, stem="学习未完成也可考试")
    db.add(ExamCandidateScope(exam_id=exam.id, candidate_id=candidate.id))
    db.commit()
    video = _upload_video(client, duration_seconds=100)
    client.post(
        f"/api/admin/learning/videos/{video['id']}/publish", headers=_admin_headers()
    )
    headers = _candidate_headers(candidate.id)

    active = client.get("/api/exams/active", headers=headers)
    started = client.post(f"/api/exams/{exam.id}/start", headers=headers)
    started_data = started.json()["data"]
    attempt_id = started_data["attempt_id"]
    submitted = client.post(
        f"/api/attempts/{attempt_id}/submit",
        headers={
            **headers,
            "X-Attempt-Session": started_data["attempt_session_credential"],
        },
        json={"submit_type": "manual"},
    )

    assert active.status_code == 200
    assert active.json()["data"][0]["id"] == exam.id
    assert started.status_code == 200
    assert submitted.status_code == 200
