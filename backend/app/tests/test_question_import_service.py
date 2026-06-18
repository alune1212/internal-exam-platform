from io import BytesIO

import pytest
from openpyxl import Workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import ImportBatch, Question, QuestionOption
from app.services import import_service
from app.services.import_service import (
    generate_failure_report,
    import_questions_from_workbook,
)
from app.services.question_service import list_questions
from app.tests.conftest import build_workbook

QUESTION_HEADERS = [
    "category_1",
    "category_2",
    "question_type",
    "stem",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "option_e",
    "option_f",
    "correct_answer",
    "analysis",
    "difficulty",
    "score",
    "status",
    "source",
    "source_no",
    "remark",
]


def test_import_questions_persists_valid_rows_and_import_batch(db: Session) -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "category_1": "制度",
                "category_2": "安全",
                "question_type": "single",
                "stem": "以下哪项是正确做法？",
                "option_a": "保管好账号",
                "option_b": "共享密码",
                "correct_answer": "A",
                "analysis": "账号不可共享。",
                "difficulty": "easy",
                "score": 2,
                "status": "active",
                "source": "handbook",
                "source_no": "Q001",
                "remark": "导入测试",
            },
            {
                "question_type": "multiple",
                "stem": "哪些属于安全要求？",
                "option_a": "定期改密",
                "option_b": "开启 MFA",
                "option_c": "外借账号",
                "correct_answer": "B,A",
                "score": 3,
                "status": "active",
            },
        ],
    )

    result = import_questions_from_workbook(db, workbook, file_name="questions.xlsx")

    questions = db.scalars(select(Question).order_by(Question.id)).all()
    options = db.scalars(
        select(QuestionOption).order_by(
            QuestionOption.question_id, QuestionOption.sort_order
        )
    ).all()
    batch = db.scalars(select(ImportBatch)).one()

    assert result.success_count == 2
    assert result.failed_count == 0
    assert result.batch_id == batch.id
    assert [question.stem for question in questions] == [
        "以下哪项是正确做法？",
        "哪些属于安全要求？",
    ]
    assert questions[0].score == 2
    assert [
        option.label for option in options if option.question_id == questions[0].id
    ] == ["A", "B"]
    assert [
        option.label
        for option in options
        if option.question_id == questions[1].id and option.is_correct
    ] == ["A", "B"]
    assert batch.import_type == "questions"
    assert batch.file_name == "questions.xlsx"
    assert batch.total_count == 2
    assert batch.success_count == 2
    assert batch.failed_count == 0
    assert batch.error_report == []


def test_import_questions_skips_invalid_rows_and_records_failures(db: Session) -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "single",
                "stem": "合法题目",
                "option_a": "正确选项",
                "option_b": "错误选项",
                "correct_answer": "A",
                "score": 1,
                "status": "active",
            },
            {
                "question_type": "single",
                "stem": "非法题目",
                "option_a": "只有 A",
                "correct_answer": "B",
                "score": 1,
                "status": "active",
            },
        ],
    )

    result = import_questions_from_workbook(db, workbook, file_name="mixed.xlsx")

    questions = db.scalars(select(Question)).all()
    batch = db.scalars(select(ImportBatch)).one()

    assert result.success_count == 1
    assert result.failed_count == 1
    assert result.failures[0].row_number == 3
    assert result.failures[0].reason == "正确答案必须存在于选项中"
    assert [question.stem for question in questions] == ["合法题目"]
    assert batch.total_count == 2
    assert batch.success_count == 1
    assert batch.failed_count == 1
    assert batch.error_report == [
        {"row_number": 3, "reason": "正确答案必须存在于选项中"}
    ]


def test_failure_report_escapes_formula_like_file_name(db: Session) -> None:
    from openpyxl import load_workbook

    db.add(
        ImportBatch(
            import_type="questions",
            file_name='=HYPERLINK("http://example.test")',
            total_count=1,
            success_count=0,
            failed_count=1,
            status="completed",
            error_report=[{"row_number": 2, "reason": "题干不能为空"}],
        )
    )
    db.commit()
    batch = db.scalars(select(ImportBatch)).one()

    workbook = load_workbook(generate_failure_report(db, batch.id), data_only=False)

    assert workbook["导入批次"].cell(3, 2).value.startswith("'=")


def test_import_upload_rejects_files_larger_than_limit() -> None:
    file_obj = BytesIO(b"abcd")

    with pytest.raises(import_service.ImportLimitError, match="不能超过 3 字节"):
        import_service.validate_upload_file_size(file_obj, max_bytes=3)

    assert file_obj.tell() == 0


def test_parse_workbook_rejects_too_many_sheets() -> None:
    workbook = Workbook()
    workbook.active.title = "题库"
    workbook.create_sheet("额外 sheet")
    file_obj = BytesIO()
    workbook.save(file_obj)
    file_obj.seek(0)

    with pytest.raises(import_service.ImportLimitError, match="不能超过 1 个工作表"):
        import_service.parse_workbook(file_obj, max_sheets=1)


def test_parse_workbook_rejects_rows_beyond_limit() -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "single",
                "stem": "第一题",
                "option_a": "A",
                "correct_answer": "A",
                "score": 1,
            },
            {
                "question_type": "single",
                "stem": "第二题",
                "option_a": "A",
                "correct_answer": "A",
                "score": 1,
            },
        ],
    )

    with pytest.raises(import_service.ImportLimitError, match="不能超过 1 行"):
        import_service.parse_workbook(workbook, max_rows=1)


def test_parse_workbook_accepts_rows_at_limit() -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "single",
                "stem": "第一题",
                "option_a": "A",
                "correct_answer": "A",
                "score": 1,
            }
        ],
    )

    parsed = import_service.parse_workbook(workbook, max_rows=1)

    assert parsed.total_count == 1


def test_import_questions_rejects_blank_required_cells(db: Session) -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "single",
                "stem": None,
                "option_a": "正确选项",
                "option_b": "错误选项",
                "correct_answer": "A",
                "score": 1,
                "status": "active",
            },
        ],
    )

    result = import_questions_from_workbook(db, workbook, file_name="blank.xlsx")

    questions = db.scalars(select(Question)).all()
    batch = db.scalars(select(ImportBatch)).one()

    assert result.success_count == 0
    assert result.failed_count == 1
    assert result.failures[0].reason == "题干不能为空"
    assert questions == []
    assert batch.error_report == [{"row_number": 2, "reason": "题干不能为空"}]


def test_import_questions_marks_judge_answer_from_true_false(db: Session) -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "judge",
                "stem": "安全生产月是每年六月。",
                "correct_answer": "true",
                "score": 1,
                "status": "active",
            },
            {
                "question_type": "judge",
                "stem": "应急预案制定后永远不用修改。",
                "correct_answer": "false",
                "score": 1,
                "status": "active",
            },
        ],
    )

    result = import_questions_from_workbook(db, workbook, file_name="judge.xlsx")

    questions = db.scalars(select(Question).order_by(Question.id)).all()
    options = db.scalars(
        select(QuestionOption).order_by(
            QuestionOption.question_id, QuestionOption.sort_order
        )
    ).all()

    assert result.success_count == 2
    assert result.failed_count == 0
    assert [
        option.label
        for option in options
        if option.question_id == questions[0].id and option.is_correct
    ] == ["A"]
    assert [
        option.label
        for option in options
        if option.question_id == questions[1].id and option.is_correct
    ] == ["B"]


def test_list_questions_returns_imported_questions_with_options(db: Session) -> None:
    workbook = build_workbook(
        QUESTION_HEADERS,
        [
            {
                "question_type": "single",
                "stem": "列表可见题目",
                "option_a": "A 选项",
                "option_b": "B 选项",
                "correct_answer": "B",
                "score": 1,
                "status": "active",
            }
        ],
    )
    import_questions_from_workbook(db, workbook, file_name="questions.xlsx")

    questions = list_questions(db)

    assert len(questions) == 1
    assert questions[0].stem == "列表可见题目"
    assert [option.label for option in questions[0].options] == ["A", "B"]
    assert [option.label for option in questions[0].options if option.is_correct] == [
        "B"
    ]
