"""admin 鉴权集成测试。"""

from collections.abc import Iterator
from datetime import UTC, datetime, timedelta
from io import BytesIO
from urllib.parse import unquote

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import ValidationError
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.config import Settings, settings
from app.core.database import Base, get_db
from app.core.security import _sign, create_session_token, verify_session_token
from app.main import create_app
from app.models import (
    AdminAuditEvent,
    Exam,
    ExamAttempt,
    ExamCandidateScope,
    ExamQuestionPool,
    ImportBatch,
    Question,
)
from app.services import exam_service
from app.tests.conftest import (
    build_workbook,
    create_candidate,
    create_exam,
    create_question_with_options,
    submit_answers,
)

PRODUCTION_PUBLIC_URL = "https://exam.example.com"


def _build_client() -> tuple[TestClient, Session]:
    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    session_local = sessionmaker(bind=engine, autoflush=False, expire_on_commit=False)
    db = session_local()
    app = create_app()

    def override_get_db() -> Iterator[Session]:
        yield db

    app.dependency_overrides[get_db] = override_get_db
    return TestClient(app), db


def test_admin_login_throttles_repeated_public_attempts(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(settings, "public_token_rate_limit_count", 2, raising=False)
    monkeypatch.setattr(
        settings, "public_token_rate_limit_window_seconds", 60, raising=False
    )
    client, _ = _build_client()

    for _ in range(2):
        resp = client.post(
            "/api/admin/login",
            json={"username": "rate-limit-admin", "password": "wrong"},
        )
        assert resp.status_code == 401

    blocked = client.post(
        "/api/admin/login",
        json={"username": "rate-limit-admin", "password": "wrong"},
    )

    assert blocked.status_code == 429


def test_admin_login_returns_token() -> None:
    client, _ = _build_client()
    resp = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["success"] is True
    assert "token" in body["data"]
    assert body["data"]["token"] != settings.admin_password


def test_admin_login_rejects_wrong_password() -> None:
    client, _ = _build_client()
    resp = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": "wrong"},
    )
    assert resp.status_code == 401
    assert "detail" in resp.json()


def test_admin_exams_requires_token() -> None:
    client, _ = _build_client()
    resp = client.get("/api/admin/exams")
    assert resp.status_code == 401


def test_admin_exams_accepts_valid_token() -> None:
    client, _ = _build_client()
    login = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    token = login.json()["data"]["token"]
    resp = client.get(
        "/api/admin/exams",
        headers={"X-Admin-Token": token},
    )
    assert resp.status_code == 200
    body = resp.json()
    assert body["success"] is True
    assert isinstance(body["data"], list)


def test_session_closure_readiness_is_admin_only_and_blocks_in_progress() -> None:
    client, db = _build_client()
    unauthorized = client.get("/api/admin/operations/session-closure-readiness")
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]
    headers = {"X-Admin-Token": token}

    ready = client.get(
        "/api/admin/operations/session-closure-readiness", headers=headers
    )
    exam = create_exam(db)
    candidate = create_candidate(db)
    now = datetime.now(UTC)
    db.add(
        ExamAttempt(
            exam_id=exam.id,
            candidate_id=candidate.id,
            status="in_progress",
            started_at=now,
            ends_at=now + timedelta(hours=1),
            duration_minutes_snapshot=60,
        )
    )
    db.commit()
    blocked = client.get(
        "/api/admin/operations/session-closure-readiness", headers=headers
    )

    assert unauthorized.status_code == 401
    assert ready.json()["data"] == {
        "ready": True,
        "in_progress_attempt_count": 0,
    }
    assert blocked.json()["data"] == {
        "ready": False,
        "in_progress_attempt_count": 1,
    }


def test_admin_exams_returns_zero_for_exam_without_question_pool() -> None:
    client, db = _build_client()
    empty_exam = create_exam(db, title="零题池考试", status="draft")
    pooled_exam = create_exam(db, title="已有题池考试", status="draft")
    question = create_question_with_options(db)
    db.add(
        ExamQuestionPool(
            exam_id=pooled_exam.id,
            question_id=question.id,
            sort_order=0,
        )
    )
    db.commit()

    login = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    token = login.json()["data"]["token"]

    response = client.get(
        "/api/admin/exams",
        headers={"X-Admin-Token": token},
    )

    assert response.status_code == 200
    exams_by_id = {item["id"]: item for item in response.json()["data"]}
    assert exams_by_id[empty_exam.id]["question_pool_count"] == 0
    assert exams_by_id[pooled_exam.id]["question_pool_count"] == 1


def test_admin_create_exam_rejects_direct_active_status() -> None:
    client, db = _build_client()
    login = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    token = login.json()["data"]["token"]

    resp = client.post(
        "/api/admin/exams",
        headers={"X-Admin-Token": token},
        json={"title": "直接上线", "duration_minutes": 60, "status": "active"},
    )

    assert resp.status_code == 422
    assert db.query(Exam).filter(Exam.title == "直接上线").first() is None


def test_admin_exams_rejects_password_as_token() -> None:
    client, _ = _build_client()
    resp = client.get(
        "/api/admin/exams",
        headers={"X-Admin-Token": settings.admin_password},
    )
    assert resp.status_code == 401


def test_admin_exams_rejects_wrong_token() -> None:
    client, _ = _build_client()
    resp = client.get(
        "/api/admin/exams",
        headers={"X-Admin-Token": "wrong"},
    )
    assert resp.status_code == 401


def test_admin_exams_rejects_expired_token(monkeypatch: pytest.MonkeyPatch) -> None:
    client, _ = _build_client()
    token = create_session_token(settings.admin_username)
    monkeypatch.setattr(settings, "admin_token_ttl_seconds", -1)

    resp = client.get(
        "/api/admin/exams",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 401


def test_named_primary_and_default_disabled_backup_operator(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(settings, "primary_operator_username", "primary")
    monkeypatch.setattr(settings, "primary_operator_password", "primary-password")
    monkeypatch.setattr(settings, "backup_operator_username", "backup")
    monkeypatch.setattr(settings, "backup_operator_password", "backup-password")
    monkeypatch.setattr(settings, "backup_operator_enabled", False)
    client, db = _build_client()

    primary = client.post(
        "/api/admin/login",
        json={"username": "primary", "password": "primary-password"},
    )
    backup = client.post(
        "/api/admin/login",
        json={"username": "backup", "password": "backup-password"},
    )

    assert primary.status_code == 200
    assert backup.status_code == 401
    events = db.query(AdminAuditEvent).order_by(AdminAuditEvent.id).all()
    assert [(event.operator_subject, event.result) for event in events] == [
        ("primary", "success"),
        ("backup", "rejected"),
    ]


def test_enabled_backup_operator_has_same_admin_access(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(settings, "primary_operator_username", "primary")
    monkeypatch.setattr(settings, "primary_operator_password", "primary-password")
    monkeypatch.setattr(settings, "backup_operator_username", "backup")
    monkeypatch.setattr(settings, "backup_operator_password", "backup-password")
    monkeypatch.setattr(settings, "backup_operator_enabled", True)
    client, _ = _build_client()

    primary_login = client.post(
        "/api/admin/login",
        json={"username": "primary", "password": "primary-password"},
    )
    login = client.post(
        "/api/admin/login",
        json={"username": "backup", "password": "backup-password"},
    )
    token = login.json()["data"]["token"]
    response = client.get("/api/admin/exams", headers={"X-Admin-Token": token})

    assert primary_login.status_code == 401
    assert login.status_code == 200
    assert response.status_code == 200


def test_switching_active_operator_invalidates_old_tokens(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    monkeypatch.setattr(settings, "primary_operator_username", "primary")
    monkeypatch.setattr(settings, "primary_operator_password", "primary-password")
    monkeypatch.setattr(settings, "backup_operator_username", "backup")
    monkeypatch.setattr(settings, "backup_operator_password", "backup-password")
    monkeypatch.setattr(settings, "backup_operator_enabled", False)
    client, _ = _build_client()

    primary_login = client.post(
        "/api/admin/login",
        json={"username": "primary", "password": "primary-password"},
    )
    primary_token = primary_login.json()["data"]["token"]
    assert (
        client.get(
            "/api/admin/exams", headers={"X-Admin-Token": primary_token}
        ).status_code
        == 200
    )

    monkeypatch.setattr(settings, "backup_operator_enabled", True)
    assert (
        client.get(
            "/api/admin/exams", headers={"X-Admin-Token": primary_token}
        ).status_code
        == 401
    )
    backup_login = client.post(
        "/api/admin/login",
        json={"username": "backup", "password": "backup-password"},
    )
    backup_token = backup_login.json()["data"]["token"]
    assert backup_login.status_code == 200
    assert (
        client.get(
            "/api/admin/exams", headers={"X-Admin-Token": backup_token}
        ).status_code
        == 200
    )

    monkeypatch.setattr(settings, "backup_operator_enabled", False)
    assert (
        client.get(
            "/api/admin/exams", headers={"X-Admin-Token": backup_token}
        ).status_code
        == 401
    )
    primary_login_after_switch = client.post(
        "/api/admin/login",
        json={"username": "primary", "password": "primary-password"},
    )
    assert primary_login_after_switch.status_code == 200


def test_admin_token_rejects_future_issued_at() -> None:
    issued_at = int((datetime.now(UTC) + timedelta(hours=1)).timestamp())
    payload = f"{settings.admin_username}.{issued_at}.nonce"
    token = f"{payload}.{_sign(payload, secret=settings.token_secret)}"

    assert (
        verify_session_token(
            token, subject=settings.admin_username, secret=settings.token_secret
        )
        is False
    )


def test_production_rejects_default_admin_password_and_token_secret() -> None:
    with pytest.raises(ValidationError):
        Settings(
            environment="production",
            cors_origins=PRODUCTION_PUBLIC_URL,
            candidate_public_base_url=PRODUCTION_PUBLIC_URL,
        )

    with pytest.raises(ValidationError):
        Settings(
            environment="production",
            admin_password="strong-password",  # noqa: S106
            token_secret="change-me-in-production",  # noqa: S106
            cors_origins=PRODUCTION_PUBLIC_URL,
            candidate_public_base_url=PRODUCTION_PUBLIC_URL,
        )


def test_production_rejects_sample_admin_password() -> None:
    with pytest.raises(ValidationError, match="ADMIN_PASSWORD"):
        Settings(
            environment="production",
            admin_password="local-dev-admin-password",  # noqa: S106
            token_secret="prod-token-secret",  # noqa: S106
            cors_origins=PRODUCTION_PUBLIC_URL,
            candidate_public_base_url=PRODUCTION_PUBLIC_URL,
        )


def test_production_rejects_sample_token_secret() -> None:
    with pytest.raises(ValidationError, match="TOKEN_SECRET"):
        Settings(
            environment="production",
            admin_password="strong-password",  # noqa: S106
            token_secret="local-dev-token-secret-change-before-production",  # noqa: S106
            cors_origins=PRODUCTION_PUBLIC_URL,
            candidate_public_base_url=PRODUCTION_PUBLIC_URL,
        )


@pytest.mark.parametrize(
    "cors_origins",
    [
        "",
        "*",
        "https://exam.example.com,*",
        "http://exam.example.com",
        "http://localhost:5173",
        "http://127.0.0.1:8080",
        "http://0.0.0.0:8080",
    ],
)
def test_production_rejects_dangerous_cors_origins(cors_origins: str) -> None:
    with pytest.raises(ValidationError, match="CORS_ORIGINS"):
        Settings(
            environment="production",
            admin_password="strong-password",  # noqa: S106
            token_secret="prod-token-secret",  # noqa: S106
            cors_origins=cors_origins,
            candidate_public_base_url=PRODUCTION_PUBLIC_URL,
            candidate_login_email_delivery_mode="smtp",
            candidate_login_email_from="noreply@example.com",
            candidate_login_smtp_host="smtp.example.com",
        )


def test_production_accepts_explicit_https_cors_origins() -> None:
    configured = Settings(
        environment="production",
        admin_password="strong-password",  # noqa: S106
        token_secret="prod-token-secret",  # noqa: S106
        cors_origins="https://exam.example.com, https://admin.example.com",
        candidate_public_base_url=PRODUCTION_PUBLIC_URL,
        candidate_login_email_delivery_mode="smtp",
        candidate_login_email_from="noreply@example.com",
        candidate_login_smtp_host="smtp.example.com",
    )

    assert configured.cors_origin_list == [
        "https://exam.example.com",
        "https://admin.example.com",
    ]


def test_admin_questions_requires_token() -> None:
    client, _ = _build_client()
    resp = client.get("/api/admin/questions")
    assert resp.status_code == 401


def test_admin_reports_requires_token() -> None:
    client, _ = _build_client()
    resp = client.get("/api/admin/reports/scores")
    assert resp.status_code == 401


def test_admin_imports_requires_token() -> None:
    client, _ = _build_client()
    resp = client.get("/api/admin/imports/templates/questions")
    assert resp.status_code == 401


def test_admin_import_failure_report_download_returns_workbook() -> None:
    client, db = _build_client()
    db.add(
        ImportBatch(
            import_type="questions",
            file_name="questions.xlsx",
            total_count=2,
            success_count=1,
            failed_count=1,
            status="completed",
            error_report=[{"row_number": 3, "reason": "题干不能为空"}],
        )
    )
    db.commit()
    batch = db.query(ImportBatch).one()
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]

    resp = client.get(
        f"/api/admin/imports/{batch.id}/failure-report",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "失败明细.xlsx" in unquote(resp.headers["content-disposition"])
    workbook = load_workbook(BytesIO(resp.content))
    assert workbook.sheetnames == ["导入批次", "失败明细"]
    meta = workbook["导入批次"]
    assert meta.cell(1, 1).value == "字段"
    assert meta.cell(1, 2).value == "值"
    assert meta.cell(2, 1).value == "导入类型"
    assert meta.cell(2, 2).value == "QUESTION IMPORT · 题库导入"
    assert meta.cell(3, 2).value == "questions.xlsx"
    assert meta.cell(4, 2).value == 2
    assert meta.cell(5, 2).value == 1
    assert meta.cell(6, 2).value == 1
    sheet = workbook["失败明细"]
    assert sheet.cell(1, 1).value == "ROW · 行号"
    assert sheet.cell(1, 2).value == "REASON · 原因"
    assert sheet.cell(2, 1).value == 3
    assert sheet.cell(2, 2).value == "题干不能为空"


def test_admin_import_failure_report_returns_empty_detail_sheet_without_failures() -> (
    None
):
    client, db = _build_client()
    db.add(
        ImportBatch(
            import_type="exam_candidates",
            file_name="名单.xlsx",
            total_count=2,
            success_count=2,
            failed_count=0,
            status="completed",
            error_report=[],
        )
    )
    db.commit()
    batch = db.query(ImportBatch).one()
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]

    resp = client.get(
        f"/api/admin/imports/{batch.id}/failure-report",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 200
    workbook = load_workbook(BytesIO(resp.content))
    assert workbook["导入批次"].cell(2, 2).value == "ROSTER IMPORT · 应考名单导入"
    assert workbook["失败明细"].max_row == 1


def test_admin_import_failure_report_returns_404_for_missing_batch() -> None:
    client, _ = _build_client()
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]

    resp = client.get(
        "/api/admin/imports/999/failure-report",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 404


def test_admin_legacy_candidate_template_is_unsupported() -> None:
    client, _ = _build_client()
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]
    legacy_template_kind = "candidate" + "s"

    resp = client.get(
        f"/api/admin/imports/templates/{legacy_template_kind}",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 404


def test_admin_question_template_download_returns_workbook() -> None:
    client, _ = _build_client()
    login = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    token = login.json()["data"]["token"]

    resp = client.get(
        "/api/admin/imports/templates/questions",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert "题库导入模板.xlsx" in unquote(resp.headers["content-disposition"])
    workbook = load_workbook(BytesIO(resp.content))
    assert workbook.active.title == "题库导入模板"
    assert workbook.active.cell(1, 1).value == "category_1"


def test_admin_report_export_returns_workbook() -> None:
    client, _ = _build_client()
    login = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    )
    token = login.json()["data"]["token"]

    resp = client.get(
        "/api/admin/reports/export",
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 200
    assert (
        resp.headers["content-type"]
        == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(resp.content))
    assert workbook.sheetnames == ["个人成绩", "题目正确率", "错题排行", "参考状态"]


def test_admin_score_report_accepts_exam_filter() -> None:
    client, db = _build_client()
    first_exam = create_exam(db, title="第一场")
    second_exam = create_exam(db, title="第二场")
    candidate = create_candidate(db, email="report@example.com")
    db.add_all(
        [
            ExamCandidateScope(
                exam_id=first_exam.id,
                candidate_id=candidate.id,
                roster_email=candidate.email,
                roster_name=candidate.name or "报告用户",
            ),
            ExamCandidateScope(
                exam_id=second_exam.id,
                candidate_id=candidate.id,
                roster_email=candidate.email,
                roster_name=candidate.name or "报告用户",
            ),
        ]
    )
    db.commit()
    create_question_with_options(db)
    first_start = exam_service.start_exam(db, first_exam.id, candidate.id)
    submit_answers(db, first_start.attempt_id, first_start.questions, ["A"])
    second_start = exam_service.start_exam(db, second_exam.id, candidate.id)
    submit_answers(db, second_start.attempt_id, second_start.questions, ["B"])
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]

    resp = client.get(
        "/api/admin/reports/scores",
        params={"exam_id": first_exam.id},
        headers={"X-Admin-Token": token},
    )

    assert resp.status_code == 200
    rows = resp.json()["data"]
    assert len(rows) == 1
    assert rows[0]["exam_title"] == "第一场"


def test_publish_rolls_back_exam_when_audit_write_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    client, db = _build_client()
    exam = create_exam(db, title="待发布", status="draft")
    candidate = create_candidate(db, email="publish@example.com")
    db.add(
        ExamCandidateScope(
            exam_id=exam.id,
            candidate_id=candidate.id,
            roster_email=candidate.email,
            roster_name=candidate.name or "发布用户",
        )
    )
    db.commit()
    create_question_with_options(db)
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]

    def fail_audit(*_args: object, **_kwargs: object) -> None:
        raise RuntimeError("audit unavailable")

    monkeypatch.setattr("app.api.exams.record_admin_event", fail_audit)
    with pytest.raises(RuntimeError, match="audit unavailable"):
        client.post(
            f"/api/admin/exams/{exam.id}/publish",
            headers={"X-Admin-Token": token},
            json={"confirmation_title": exam.title},
        )
    db.expire_all()
    persisted_exam = db.get(Exam, exam.id)
    assert persisted_exam is not None
    assert persisted_exam.status == "draft"
    assert db.query(ExamQuestionPool).filter_by(exam_id=exam.id).count() == 0


def test_question_import_rolls_back_questions_when_audit_write_fails(
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    client, db = _build_client()
    token = client.post(
        "/api/admin/login",
        json={"username": "admin", "password": settings.admin_password},
    ).json()["data"]["token"]
    workbook = build_workbook(
        ["question_type", "stem", "option_a", "option_b", "correct_answer", "score"],
        [
            {
                "question_type": "single",
                "stem": "导入后应回滚",
                "option_a": "正确",
                "option_b": "错误",
                "correct_answer": "A",
                "score": 1,
            }
        ],
    )

    def fail_audit(*_args: object, **_kwargs: object) -> None:
        raise RuntimeError("audit unavailable")

    monkeypatch.setattr("app.api.questions.record_admin_event", fail_audit)
    with pytest.raises(RuntimeError, match="audit unavailable"):
        client.post(
            "/api/admin/questions/import",
            headers={"X-Admin-Token": token},
            files={
                "file": (
                    "questions.xlsx",
                    workbook.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
    db.expire_all()
    assert db.query(Question).count() == 0
    assert db.query(ImportBatch).filter_by(import_type="questions").count() == 0
