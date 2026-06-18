from sqlalchemy.orm import Session

from app.models import ExamCandidateScope
from app.services import exam_service, report_service
from app.tests.conftest import (
    create_candidate,
    create_exam,
    create_question_with_options,
    submit_answers,
)


def _setup_exam_with_candidates(db: Session):
    """创建一场考试、两个考生、两道题，返回 (exam, c1, c2, start1, start2)。"""
    exam = create_exam(db, title="测试考试")
    c1 = create_candidate(db, name="张三", employee_no="E001", department="技术部")
    c2 = create_candidate(db, name="李四", employee_no="E002", department="产品部")
    db.add_all(
        [
            ExamCandidateScope(exam_id=exam.id, candidate_id=c1.id),
            ExamCandidateScope(exam_id=exam.id, candidate_id=c2.id),
        ]
    )
    db.commit()
    create_question_with_options(db, stem="题目A", score=5)
    create_question_with_options(db, stem="题目B", score=5)

    r1 = exam_service.start_exam(db, exam.id, c1.id)
    r2 = exam_service.start_exam(db, exam.id, c2.id)
    return exam, c1, c2, r1, r2


# --- 成绩报表 ---


def test_score_report_returns_submitted_attempts(db: Session) -> None:
    exam, c1, c2, r1, r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    submit_answers(db, r2.attempt_id, r2.questions, ["B", "A"])

    report = report_service.get_score_report(db)

    assert len(report) == 2
    # 按分数降序
    assert report[0].candidate_name == "张三"
    assert report[0].score == 10
    assert report[1].candidate_name == "李四"
    assert report[1].score == 5


def test_score_report_excludes_in_progress(db: Session) -> None:
    exam, c1, c2, r1, r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])

    report = report_service.get_score_report(db)
    assert len(report) == 1


def test_score_report_uses_latest_submitted_attempt(db: Session) -> None:
    exam, c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    exam_service.create_retake_grant(db, exam.id, c1.id)
    retake = exam_service.start_exam(db, exam.id, c1.id)
    submit_answers(db, retake.attempt_id, retake.questions, ["B", "B"])

    report = report_service.get_score_report(db)
    row = next(item for item in report if item.candidate_name == "张三")

    assert row.score == 0


def test_reports_filter_by_exam_id(db: Session) -> None:
    exam, c1, _c2, r1, r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    submit_answers(db, r2.attempt_id, r2.questions, ["B", "B"])

    second_exam = create_exam(db, title="第二场考试")
    db.add(ExamCandidateScope(exam_id=second_exam.id, candidate_id=c1.id))
    db.commit()
    second_start = exam_service.start_exam(db, second_exam.id, c1.id)
    submit_answers(db, second_start.attempt_id, second_start.questions, ["B", "B"])

    first_scores = report_service.get_score_report(db, exam_id=exam.id)
    second_scores = report_service.get_score_report(db, exam_id=second_exam.id)
    second_accuracy = report_service.get_question_accuracy(db, exam_id=second_exam.id)
    first_wrong = report_service.get_wrong_questions(db, exam_id=exam.id)

    assert {row.exam_title for row in first_scores} == {"测试考试"}
    assert {row.candidate_name for row in first_scores} == {"张三", "李四"}
    assert [row.exam_title for row in second_scores] == ["第二场考试"]
    assert second_scores[0].score == 0
    assert all(row.total_count == 1 for row in second_accuracy)
    assert all(row.correct_count == 0 for row in second_accuracy)
    assert {row.stem for row in first_wrong} == {"题目A", "题目B"}


# --- 题目正确率 ---


def test_question_accuracy_calculates_rate(db: Session) -> None:
    exam, c1, c2, r1, r2 = _setup_exam_with_candidates(db)
    # 两人都答对第1题，都答错第2题
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "B"])
    submit_answers(db, r2.attempt_id, r2.questions, ["A", "B"])

    report = report_service.get_question_accuracy(db)

    assert len(report) == 2
    # 按正确率升序
    assert report[0].accuracy_rate == 0.0  # 题目B
    assert report[0].total_count == 2
    assert report[0].correct_count == 0
    assert report[1].accuracy_rate == 1.0  # 题目A
    assert report[1].correct_count == 2


def test_question_accuracy_uses_latest_submitted_attempt(db: Session) -> None:
    exam, c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["B", "B"])
    exam_service.create_retake_grant(db, exam.id, c1.id)
    retake = exam_service.start_exam(db, exam.id, c1.id)
    submit_answers(db, retake.attempt_id, retake.questions, ["A", "A"])

    report = report_service.get_question_accuracy(db)

    assert all(row.accuracy_rate == 1.0 for row in report)


# --- 错题统计 ---


def test_wrong_questions_counts_failures(db: Session) -> None:
    exam, c1, c2, r1, r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "B"])
    submit_answers(db, r2.attempt_id, r2.questions, ["A", "B"])

    report = report_service.get_wrong_questions(db)

    assert len(report) == 1
    assert report[0].stem == "题目B"
    assert report[0].wrong_count == 2


def test_wrong_questions_uses_latest_submitted_attempt_after_retake(
    db: Session,
) -> None:
    exam, c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "B"])
    exam_service.create_retake_grant(db, exam.id, c1.id)
    retake = exam_service.start_exam(db, exam.id, c1.id)
    submit_answers(db, retake.attempt_id, retake.questions, ["A", "A"])

    report = report_service.get_wrong_questions(db)

    assert all(row.stem != "题目B" for row in report)


# --- 缺考人员 ---


def test_absent_candidates_splits_not_started_and_in_progress(db: Session) -> None:
    exam, c1, c2, r1, r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    # c2 已开考但未提交，不应再算作未开始缺考。

    c3 = create_candidate(db, name="王五", employee_no="E003")
    db.add(ExamCandidateScope(exam_id=exam.id, candidate_id=c3.id))
    db.commit()

    report = report_service.get_absent_candidates(db, exam_id=exam.id)

    names = [r.name for r in report]
    assert "王五" in names
    assert "张三" not in names
    assert "李四" not in names

    in_progress = report_service.get_absent_candidates(
        db, exam_id=exam.id, status="in_progress"
    )
    assert [r.name for r in in_progress] == ["李四"]
    submitted = report_service.get_absent_candidates(
        db, exam_id=exam.id, status="submitted"
    )
    assert [r.name for r in submitted] == ["张三"]


def test_absent_candidates_excludes_non_should_attend(db: Session) -> None:
    create_candidate(db, name="不需要参加", employee_no="E010", should_attend=False)

    report = report_service.get_absent_candidates(db)
    names = [r.name for r in report]
    assert "不需要参加" not in names


def test_absent_candidates_global_status_filter(db: Session) -> None:
    exam, c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    c3 = create_candidate(db, name="王五", employee_no="E003")
    db.add(ExamCandidateScope(exam_id=exam.id, candidate_id=c3.id))
    db.commit()

    assert [r.name for r in report_service.get_absent_candidates(db)] == ["王五"]
    assert [
        r.name for r in report_service.get_absent_candidates(db, status="in_progress")
    ] == ["李四"]
    assert [
        r.name for r in report_service.get_absent_candidates(db, status="submitted")
    ] == ["张三"]


def test_report_workbook_contains_all_report_sheets(db: Session) -> None:
    from openpyxl import load_workbook

    exam, _c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "B"])

    workbook_stream = report_service.generate_report_workbook(db)
    workbook = load_workbook(workbook_stream)

    assert workbook.sheetnames == ["成绩报表", "题目正确率", "错题统计", "参考状态"]
    assert [cell.value for cell in workbook["成绩报表"][1]] == [
        "姓名",
        "员工号",
        "部门",
        "考试",
        "得分",
        "总分",
        "提交时间",
    ]
    assert workbook["成绩报表"].cell(2, 1).value == "张三"
    assert [cell.value for cell in workbook["参考状态"][1]] == [
        "考生ID",
        "姓名",
        "员工号",
        "部门",
        "考试分组",
        "参考状态",
    ]


def test_report_workbook_escapes_formula_like_text(db: Session) -> None:
    from openpyxl import load_workbook

    exam = create_exam(db, title="=cmd")
    candidate = create_candidate(db, name='=HYPERLINK("http://example.test")')
    db.add(ExamCandidateScope(exam_id=exam.id, candidate_id=candidate.id))
    db.commit()
    create_question_with_options(db, stem="+SUM(1,1)")
    start = exam_service.start_exam(db, exam.id, candidate.id)
    submit_answers(db, start.attempt_id, start.questions, ["B"])

    workbook_stream = report_service.generate_report_workbook(db, exam_id=exam.id)
    workbook = load_workbook(workbook_stream, data_only=False)

    assert workbook["成绩报表"].cell(2, 1).value.startswith("'=")
    assert workbook["成绩报表"].cell(2, 4).value == "'=cmd"
    assert workbook["题目正确率"].cell(2, 2).value == "'+SUM(1,1)"
    assert workbook["错题统计"].cell(2, 2).value == "'+SUM(1,1)"


def test_report_workbook_filters_by_exam_id(db: Session) -> None:
    from openpyxl import load_workbook

    exam, c1, _c2, r1, _r2 = _setup_exam_with_candidates(db)
    submit_answers(db, r1.attempt_id, r1.questions, ["A", "A"])
    second_exam = create_exam(db, title="第二场考试")
    db.add(ExamCandidateScope(exam_id=second_exam.id, candidate_id=c1.id))
    db.commit()
    second_start = exam_service.start_exam(db, second_exam.id, c1.id)
    submit_answers(db, second_start.attempt_id, second_start.questions, ["B", "B"])

    workbook_stream = report_service.generate_report_workbook(db, exam_id=exam.id)
    workbook = load_workbook(workbook_stream)

    assert workbook["成绩报表"].cell(2, 4).value == "测试考试"
    assert workbook["成绩报表"].cell(3, 4).value is None
