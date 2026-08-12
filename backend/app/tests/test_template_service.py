from io import BytesIO

from openpyxl import load_workbook

from app.services.template_service import (
    EXAM_ROSTER_HEADERS,
    generate_exam_roster_template,
    generate_question_template,
)


def test_question_template_has_correct_headers() -> None:
    wb = load_workbook(generate_question_template())
    ws = wb.active
    assert ws.title == "题库导入模板"
    assert [cell.value for cell in ws[1]] == [
        "category_1",
        "category_2",
        "question_type",
        "stem",
        "option_a",
        "option_b",
        "option_c",
        "option_d",
        "option_e",
        "option_f",
        "correct_answer",
        "analysis",
        "difficulty",
        "score",
        "status",
        "source",
        "source_no",
        "remark",
    ]


def test_exam_roster_template_has_reduced_contract() -> None:
    wb = load_workbook(generate_exam_roster_template())
    ws = wb.active
    assert ws.title == "应考名单导入模板"
    assert [cell.value for cell in ws[1]] == EXAM_ROSTER_HEADERS
    assert ws.max_row == 2
    assert ws.cell(2, 1).value == "zhangsan@example.com"
    assert ws.cell(2, 2).value == "张三"


def test_templates_return_bytesio() -> None:
    assert isinstance(generate_question_template(), BytesIO)
    assert isinstance(generate_exam_roster_template(), BytesIO)
