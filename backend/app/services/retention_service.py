from __future__ import annotations

import hashlib
import json
import re
import zipfile
from datetime import UTC, datetime, time, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from sqlalchemy import func
from sqlalchemy.orm import Session, selectinload

from app.core.config import settings
from app.core.exceptions import DomainError
from app.core.time import to_utc
from app.models import (
    AdminAuditEvent,
    Exam,
    ExamAttempt,
    ExamAttemptAnswer,
    ExamAttemptQuestion,
    ExamCandidateScope,
    ExamQuestionPool,
    ExamRetakeGrant,
)
from app.ops.internal_backup import BackupValidationError, validate_backup
from app.schemas.operations import (
    RetentionArchiveRead,
    RetentionDeleteRead,
    RetentionExamPreview,
    RetentionPreviewRead,
)
from app.services.audit_service import record_admin_event
from app.services.operational_lock_service import assert_admin_mutation_allowed

RETENTION_MONTHS = 12
RETENTION_DAYS = 365
ARTIFACT_ID_PATTERN = re.compile(r"^retention-[0-9]{8}t[0-9]{6}z-[0-9a-f]{12}$")


class RetentionSafeguardError(DomainError):
    status_code = 409


def _latest(*values: datetime | None) -> datetime:
    present = [to_utc(value) for value in values if value is not None]
    return max(present) if present else datetime.min.replace(tzinfo=UTC)


def _fingerprint(cutoff_at: datetime, exams: list[RetentionExamPreview]) -> str:
    payload = {
        "cutoff_at": cutoff_at.isoformat(),
        "exams": [row.model_dump(mode="json") for row in exams],
    }
    return hashlib.sha256(
        json.dumps(payload, ensure_ascii=False, sort_keys=True).encode("utf-8")
    ).hexdigest()


def preview_retention(
    db: Session, *, now: datetime | None = None
) -> RetentionPreviewRead:
    generated_at = to_utc(now or datetime.now(UTC))
    cutoff_at = datetime.combine(
        (generated_at - timedelta(days=RETENTION_DAYS)).date(), time.min, UTC
    )
    rows: list[RetentionExamPreview] = []
    exams = db.query(Exam).order_by(Exam.id).all()
    for exam in exams:
        attempts = db.query(ExamAttempt).filter(ExamAttempt.exam_id == exam.id).all()
        attempt_ids = [attempt.id for attempt in attempts]
        attempt_question_count = (
            db.query(func.count(ExamAttemptQuestion.id))
            .join(ExamAttempt, ExamAttemptQuestion.attempt_id == ExamAttempt.id)
            .filter(ExamAttempt.exam_id == exam.id)
            .scalar()
            or 0
        )
        answer_count = (
            db.query(func.count(ExamAttemptAnswer.id))
            .join(
                ExamAttemptQuestion,
                ExamAttemptAnswer.attempt_question_id == ExamAttemptQuestion.id,
            )
            .join(ExamAttempt, ExamAttemptQuestion.attempt_id == ExamAttempt.id)
            .filter(ExamAttempt.exam_id == exam.id)
            .scalar()
            or 0
        )
        final_activity_at = _latest(
            exam.created_at,
            exam.updated_at,
            *(
                value
                for attempt in attempts
                for value in (
                    attempt.created_at,
                    attempt.updated_at,
                    attempt.submitted_at,
                    attempt.voided_at,
                )
            ),
        )
        reasons: list[str] = []
        if exam.status != "archived":
            reasons.append("考试尚未归档")
        if any(attempt.status == "in_progress" for attempt in attempts):
            reasons.append("仍有进行中的正式考试记录")
        if final_activity_at > cutoff_at:
            reasons.append("最终活动距今未满 12 个月")
        candidate_ids = {
            row[0]
            for row in db.query(ExamCandidateScope.candidate_id)
            .filter(ExamCandidateScope.exam_id == exam.id)
            .all()
        } | {attempt.candidate_id for attempt in attempts}
        audit_evidence_count = sum(
            1
            for event in db.query(AdminAuditEvent).all()
            if (event.target_type == "exam" and event.target_id == str(exam.id))
            or event.metadata_json.get("exam_id") == exam.id
        )
        rows.append(
            RetentionExamPreview(
                exam_id=exam.id,
                title=exam.title,
                final_activity_at=final_activity_at,
                eligible=not reasons,
                reasons=reasons or ["符合 12 个月归档删除条件"],
                attempt_count=len(attempt_ids),
                attempt_question_count=int(attempt_question_count),
                answer_count=int(answer_count),
                roster_count=db.query(ExamCandidateScope)
                .filter(ExamCandidateScope.exam_id == exam.id)
                .count(),
                retake_grant_count=db.query(ExamRetakeGrant)
                .filter(ExamRetakeGrant.exam_id == exam.id)
                .count(),
                frozen_pool_count=db.query(ExamQuestionPool)
                .filter(ExamQuestionPool.exam_id == exam.id)
                .count(),
                protected_candidate_count=len(candidate_ids),
                audit_evidence_count=audit_evidence_count,
            )
        )
    return RetentionPreviewRead(
        generated_at=generated_at,
        cutoff_at=cutoff_at,
        retention_months=RETENTION_MONTHS,
        fingerprint=_fingerprint(cutoff_at, rows),
        exams=rows,
    )


def _eligible_rows(
    preview: RetentionPreviewRead, exam_ids: list[int], fingerprint: str
) -> list[RetentionExamPreview]:
    normalized_ids = sorted(set(exam_ids))
    if not normalized_ids or normalized_ids != sorted(exam_ids):
        raise RetentionSafeguardError("必须提供非空、唯一且有序的考试 ID。")
    if preview.fingerprint != fingerprint:
        raise RetentionSafeguardError("保留预览已过期，请重新预览。")
    rows_by_id = {row.exam_id: row for row in preview.exams}
    if any(
        exam_id not in rows_by_id or not rows_by_id[exam_id].eligible
        for exam_id in normalized_ids
    ):
        raise RetentionSafeguardError("所选考试包含不符合保留删除条件的记录。")
    return [rows_by_id[exam_id] for exam_id in normalized_ids]


def _archive_payload(db: Session, exam_ids: list[int]) -> dict[str, Any]:
    exams = (
        db.query(Exam)
        .options(
            selectinload(Exam.attempts)
            .selectinload(ExamAttempt.questions)
            .selectinload(ExamAttemptQuestion.answer)
        )
        .filter(Exam.id.in_(exam_ids))
        .order_by(Exam.id)
        .all()
    )
    return {
        "exams": [
            {
                "id": exam.id,
                "title": exam.title,
                "status": exam.status,
                "available_from": exam.available_from,
                "available_until": exam.available_until,
                "attempts": [
                    {
                        "id": attempt.id,
                        "candidate_id": attempt.candidate_id,
                        "status": attempt.status,
                        "started_at": attempt.started_at,
                        "submitted_at": attempt.submitted_at,
                        "score": str(attempt.score),
                        "total_score": str(attempt.total_score),
                        "questions": [
                            {
                                "id": question.id,
                                "stem_snapshot": question.stem_snapshot,
                                "options_snapshot": question.options_snapshot,
                                "correct_answer_snapshot": question.correct_answer_snapshot,
                                "analysis_snapshot": question.analysis_snapshot,
                                "selected_answer": question.answer.selected_answer
                                if question.answer
                                else None,
                                "is_correct": question.answer.is_correct
                                if question.answer
                                else None,
                            }
                            for question in attempt.questions
                        ],
                    }
                    for attempt in attempt_sort(exam.attempts)
                ],
            }
            for exam in exams
        ]
    }


def attempt_sort(attempts: list[ExamAttempt]) -> list[ExamAttempt]:
    return sorted(attempts, key=lambda attempt: attempt.id)


def _json_bytes(payload: object) -> bytes:
    return (
        json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True, default=str)
        + "\n"
    ).encode("utf-8")


def _workbook_bytes(payload: dict[str, Any]) -> bytes:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "考试归档"
    summary.append(["考试ID", "考试标题", "状态", "考试记录数"])
    attempts = workbook.create_sheet("考试记录")
    attempts.append(["考试ID", "记录ID", "考试人ID", "状态", "得分", "总分"])
    answers = workbook.create_sheet("作答快照")
    answers.append(["记录ID", "题目快照ID", "题干", "所选答案", "正确答案", "是否正确"])
    for exam in payload["exams"]:
        summary.append(
            [exam["id"], exam["title"], exam["status"], len(exam["attempts"])]
        )
        for attempt in exam["attempts"]:
            attempts.append(
                [
                    exam["id"],
                    attempt["id"],
                    attempt["candidate_id"],
                    attempt["status"],
                    attempt["score"],
                    attempt["total_score"],
                ]
            )
            for question in attempt["questions"]:
                answers.append(
                    [
                        attempt["id"],
                        question["id"],
                        question["stem_snapshot"],
                        question["selected_answer"],
                        question["correct_answer_snapshot"],
                        question["is_correct"],
                    ]
                )
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()


def create_retention_archive(
    db: Session,
    *,
    exam_ids: list[int],
    preview_fingerprint: str,
    operator_subject: str,
    now: datetime | None = None,
) -> RetentionArchiveRead:
    assert_admin_mutation_allowed(db)
    created_at = to_utc(now or datetime.now(UTC))
    preview = preview_retention(db, now=created_at)
    rows = _eligible_rows(preview, exam_ids, preview_fingerprint)
    payload = _archive_payload(db, [row.exam_id for row in rows])
    json_content = _json_bytes(payload)
    workbook_content = _workbook_bytes(payload)
    artifact_id = (
        f"retention-{created_at.strftime('%Y%m%dt%H%M%Sz').lower()}-"
        f"{preview.fingerprint[:12]}"
    )
    manifest = {
        "schema_version": 1,
        "artifact_id": artifact_id,
        "created_at": created_at.isoformat(),
        "operator_subject": operator_subject,
        "exam_ids": [row.exam_id for row in rows],
        "preview_fingerprint": preview.fingerprint,
        "files": {
            "archive.json": hashlib.sha256(json_content).hexdigest(),
            "archive.xlsx": hashlib.sha256(workbook_content).hexdigest(),
        },
    }
    manifest_content = _json_bytes(manifest)
    archive_dir = Path(settings.lifecycle_archive_dir)
    archive_dir.mkdir(parents=True, exist_ok=True)
    archive_path = archive_dir / f"{artifact_id}.zip"
    temporary_path = archive_dir / f".{artifact_id}.tmp"
    with zipfile.ZipFile(
        temporary_path, "w", compression=zipfile.ZIP_DEFLATED
    ) as bundle:
        bundle.writestr("archive.json", json_content)
        bundle.writestr("archive.xlsx", workbook_content)
        bundle.writestr("manifest.json", manifest_content)
    temporary_path.replace(archive_path)
    archive_sha256 = hashlib.sha256(archive_path.read_bytes()).hexdigest()
    (archive_dir / f"{artifact_id}.manifest.json").write_bytes(manifest_content)
    (archive_dir / f"{artifact_id}.sha256").write_text(
        f"{archive_sha256}  {archive_path.name}\n", encoding="ascii"
    )
    record_admin_event(
        db,
        operator_subject=operator_subject,
        action="retention_archive_created",
        target_type="exam_set",
        target_id=",".join(str(row.exam_id) for row in rows),
        metadata={"archive_ref": artifact_id, "count": len(rows)},
    )
    db.commit()
    return RetentionArchiveRead(
        artifact_id=artifact_id,
        created_at=created_at,
        exam_ids=[row.exam_id for row in rows],
        preview_fingerprint=preview.fingerprint,
        archive_sha256=archive_sha256,
    )


def _load_archive_manifest(artifact_id: str) -> dict[str, object]:
    if ARTIFACT_ID_PATTERN.fullmatch(artifact_id) is None:
        raise RetentionSafeguardError("归档产物标识无效。")
    archive_dir = Path(settings.lifecycle_archive_dir)
    archive_path = archive_dir / f"{artifact_id}.zip"
    manifest_path = archive_dir / f"{artifact_id}.manifest.json"
    checksum_path = archive_dir / f"{artifact_id}.sha256"
    if (
        not archive_path.is_file()
        or not manifest_path.is_file()
        or not checksum_path.is_file()
    ):
        raise RetentionSafeguardError("归档产物不完整。")
    expected = checksum_path.read_text(encoding="ascii").split("  ", 1)[0]
    if hashlib.sha256(archive_path.read_bytes()).hexdigest() != expected:
        raise RetentionSafeguardError("归档产物 checksum 校验失败。")
    return json.loads(manifest_path.read_text(encoding="utf-8"))


def delete_retained_exams(
    db: Session,
    *,
    exam_ids: list[int],
    preview_fingerprint: str,
    archive_id: str,
    backup_id: str,
    confirmation: str,
    operator_subject: str,
    now: datetime | None = None,
) -> RetentionDeleteRead:
    assert_admin_mutation_allowed(db)
    deleted_at = to_utc(now or datetime.now(UTC))
    preview = preview_retention(db, now=deleted_at)
    rows = _eligible_rows(preview, exam_ids, preview_fingerprint)
    normalized_ids = [row.exam_id for row in rows]
    expected_confirmation = f"DELETE EXAMS {','.join(map(str, normalized_ids))}"
    if confirmation != expected_confirmation:
        raise RetentionSafeguardError("删除确认文本不匹配。")
    manifest = _load_archive_manifest(archive_id)
    if (
        manifest.get("exam_ids") != normalized_ids
        or manifest.get("preview_fingerprint") != preview.fingerprint
    ):
        raise RetentionSafeguardError("归档产物与当前预览或所选考试不匹配。")
    try:
        backup_manifest = validate_backup(Path(settings.backup_storage_dir) / backup_id)
    except BackupValidationError as exc:
        raise RetentionSafeguardError("配对备份未通过校验。") from exc
    archive_created_at = datetime.fromisoformat(str(manifest["created_at"]))
    backup_created_at = datetime.fromisoformat(str(backup_manifest["created_at"]))
    if to_utc(backup_created_at) < to_utc(archive_created_at):
        raise RetentionSafeguardError("配对备份早于归档产物，请重新创建并验证备份。")

    attempt_ids = [
        row[0]
        for row in db.query(ExamAttempt.id)
        .filter(ExamAttempt.exam_id.in_(normalized_ids))
        .all()
    ]
    if attempt_ids:
        question_ids = [
            row[0]
            for row in db.query(ExamAttemptQuestion.id)
            .filter(ExamAttemptQuestion.attempt_id.in_(attempt_ids))
            .all()
        ]
        if question_ids:
            db.query(ExamAttemptAnswer).filter(
                ExamAttemptAnswer.attempt_question_id.in_(question_ids)
            ).delete(synchronize_session=False)
        db.query(ExamAttemptQuestion).filter(
            ExamAttemptQuestion.attempt_id.in_(attempt_ids)
        ).delete(synchronize_session=False)
    db.query(ExamRetakeGrant).filter(
        ExamRetakeGrant.exam_id.in_(normalized_ids)
    ).delete(synchronize_session=False)
    db.query(ExamCandidateScope).filter(
        ExamCandidateScope.exam_id.in_(normalized_ids)
    ).delete(synchronize_session=False)
    db.query(ExamQuestionPool).filter(
        ExamQuestionPool.exam_id.in_(normalized_ids)
    ).delete(synchronize_session=False)
    db.query(ExamAttempt).filter(ExamAttempt.exam_id.in_(normalized_ids)).delete(
        synchronize_session=False
    )
    db.query(Exam).filter(Exam.id.in_(normalized_ids)).delete(synchronize_session=False)
    protected_candidate_count = sum(row.protected_candidate_count for row in rows)
    record_admin_event(
        db,
        operator_subject=operator_subject,
        action="retention_deleted",
        target_type="exam_set",
        target_id=",".join(map(str, normalized_ids)),
        metadata={
            "archive_ref": archive_id,
            "backup_ref": backup_id,
            "deleted_count": len(normalized_ids),
        },
    )
    db.commit()
    db.expire_all()
    return RetentionDeleteRead(
        deleted_exam_ids=normalized_ids,
        deleted_attempt_count=len(attempt_ids),
        protected_candidate_count=protected_candidate_count,
        archive_id=archive_id,
        backup_id=backup_id,
    )
