from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

QUESTION_HEADERS = [
    "category_1",
    "category_2",
    "question_type",
    "stem",
    "option_a",
    "option_b",
    "option_c",
    "option_d",
    "option_e",
    "option_f",
    "correct_answer",
    "analysis",
    "difficulty",
    "score",
    "status",
    "source",
    "source_no",
    "remark",
]

QUESTION_EXAMPLES = [
    [
        "安全知识",
        "消防",
        "single",
        "灭火器的有效射程是多少米？",
        "3-5",
        "5-8",
        "8-10",
        "10-15",
        None,
        None,
        "A",
        "灭火器有效射程一般为3-5米",
        "简单",
        2,
        "active",
        None,
        None,
        None,
    ],
    [
        "安全知识",
        "消防",
        "multiple",
        "以下哪些属于灭火的基本方法？",
        "隔离法",
        "窒息法",
        "冷却法",
        "抑制法",
        None,
        None,
        "A,B,C,D",
        "四种均为灭火基本方法",
        "中等",
        2,
        "active",
        None,
        None,
        None,
    ],
]

# The standalone personnel import was retired.  The only supported candidate
# workbook is a draft exam roster keyed by normalized email.  Keep the optional
# organization columns in the template so operators can prepare the frozen
# formal identity without touching the global account profile.
EXAM_ROSTER_HEADERS = [
    "email",
    "candidate_name",
    "department",
    "position",
    "exam_group",
    "remark",
]

EXAM_ROSTER_EXAMPLES = [
    [
        "zhangsan@example.com",
        "张三",
        "综合管理部",
        "工程师",
        "A组",
        None,
    ],
]

HEADER_FONT = Font(bold=True)


def _build_workbook(
    sheet_name: str,
    headers: list[str],
    examples: list[list],
) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(1, col_idx, header)
        cell.font = HEADER_FONT

    # example rows
    for row_idx, row_data in enumerate(examples, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            if value is not None:
                ws.cell(row_idx, col_idx, value)

    # auto column width
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(header)
        for row_idx in range(2, len(examples) + 2):
            cell_value = ws.cell(row_idx, col_idx).value
            if cell_value is not None:
                max_len = max(max_len, len(str(cell_value)))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def generate_question_template() -> BytesIO:
    return _build_workbook("题库导入模板", QUESTION_HEADERS, QUESTION_EXAMPLES)


def generate_exam_roster_template() -> BytesIO:
    return _build_workbook(
        "应考名单导入模板", EXAM_ROSTER_HEADERS, EXAM_ROSTER_EXAMPLES
    )
