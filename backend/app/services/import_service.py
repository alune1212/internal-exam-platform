from contextlib import suppress
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from pydantic import EmailStr, TypeAdapter, ValidationError
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.core.config import settings
from app.core.exceptions import DomainError
from app.models import (
    Candidate,
    Exam,
    ExamCandidateScope,
    ImportBatch,
    Question,
    QuestionOption,
)
from app.schemas.candidate import normalize_email
from app.schemas.question import ImportFailure, QuestionImportResult
from app.services.exam_errors import ExamFrozenError, ExamNotFoundError
from app.services.excel_security import escape_excel_cell
from app.services.operational_lock_service import assert_admin_mutation_allowed
from app.services.scoring_service import normalize_answer_set

OPTION_LABELS = ("A", "B", "C", "D", "E", "F")
VALID_QUESTION_TYPES = {"single", "multiple", "judge"}
VALID_STATUSES = {"active", "inactive"}
DEFAULT_STATUS = "active"
EMAIL_ADAPTER = TypeAdapter(EmailStr)
JUDGE_OPTIONS = [("A", "正确"), ("B", "错误")]
JUDGE_ANSWER_MAP = {"true": "A", "false": "B"}
IMPORT_TYPE_LABELS = {
    "questions": "QUESTION IMPORT · 题库导入",
    "exam_candidates": "ROSTER IMPORT · 应考名单导入",
}

ROSTER_REQUIRED_HEADERS = {"email", "candidate_name"}
ROSTER_OPTIONAL_HEADERS = {"department", "position", "exam_group", "remark"}
ROSTER_TEXT_MAX_LENGTH = 100
ROSTER_REMARK_MAX_LENGTH = 2000


class ImportBatchNotFoundError(DomainError):
    status_code = 404

    def __init__(self, batch_id: int) -> None:
        super().__init__(f"导入批次 #{batch_id} 不存在")


class ImportLimitError(DomainError):
    status_code = 413


@dataclass(frozen=True)
class ParsedWorkbook:
    rows: list[dict[str, Any]]
    total_count: int
    headers: list[str]


def validate_upload_file_size(file_obj: Any, *, max_bytes: int | None = None) -> None:
    limit = max_bytes or settings.import_max_upload_bytes
    try:
        file_obj.seek(0, 2)
        size = file_obj.tell()
        file_obj.seek(0)
    except (AttributeError, OSError):
        return

    if size > limit:
        raise ImportLimitError(f"导入文件大小不能超过 {limit} 字节")


def parse_workbook(
    file_obj: Any, *, max_rows: int | None = None, max_sheets: int | None = None
) -> ParsedWorkbook:
    row_limit = max_rows or settings.import_max_rows
    sheet_limit = max_sheets or settings.import_max_sheets
    with suppress(AttributeError, OSError):
        file_obj.seek(0)
    workbook = load_workbook(file_obj, read_only=True, data_only=True)
    try:
        if len(workbook.worksheets) > sheet_limit:
            raise ImportLimitError(f"导入文件不能超过 {sheet_limit} 个工作表")

        sheet = workbook.active
        it = sheet.iter_rows(values_only=True)
        headers_row = next(it, None)
        if headers_row is None:
            return ParsedWorkbook(rows=[], total_count=0, headers=[])

        # Workbook contracts are case-insensitive and tolerate surrounding
        # whitespace.  Normalize once here so every importer reads the same
        # canonical row keys rather than validating one spelling and looking
        # up another.
        headers = [
            str(cell).strip().lower() if cell is not None else ""
            for cell in headers_row
        ]
        parsed_rows = []
        for row_number, row in enumerate(it, start=1):
            if row_number > row_limit:
                raise ImportLimitError(f"导入数据行数不能超过 {row_limit} 行")
            parsed_rows.append(
                {headers[i]: v for i, v in enumerate(row) if i < len(headers)}
            )
        return ParsedWorkbook(
            rows=parsed_rows, total_count=len(parsed_rows), headers=headers
        )
    finally:
        workbook.close()


def import_questions_from_workbook(
    db: Session,
    file_obj: Any,
    file_name: str,
    *,
    commit: bool = True,
) -> QuestionImportResult:
    assert_admin_mutation_allowed(db)
    validate_upload_file_size(file_obj)
    parsed = parse_workbook(file_obj)
    failures: list[ImportFailure] = []
    imported_questions: list[Question] = []

    for row_number, row in enumerate(parsed.rows, start=2):
        reason = validate_question_import_row(row)
        if reason:
            failures.append(ImportFailure(row_number=row_number, reason=reason))
            continue
        imported_questions.append(_build_question(row))

    db.add_all(imported_questions)
    batch = ImportBatch(
        import_type="questions",
        file_name=file_name,
        total_count=parsed.total_count,
        success_count=len(imported_questions),
        failed_count=len(failures),
        status="completed",
        error_report=[failure.model_dump() for failure in failures],
    )
    db.add(batch)
    db.flush()
    if commit:
        db.commit()

    return QuestionImportResult(
        batch_id=batch.id,
        success_count=len(imported_questions),
        failed_count=len(failures),
        failures=failures,
    )


def validate_exam_roster_headers(headers: list[str]) -> str | None:
    """Validate the reduced workbook contract before touching the database."""

    normalized = {header.strip().lower() for header in headers if header.strip()}
    missing = ROSTER_REQUIRED_HEADERS.difference(normalized)
    if missing:
        return "应考名单必须包含 email 和 candidate_name 列"
    unsupported = normalized.difference(
        ROSTER_REQUIRED_HEADERS | ROSTER_OPTIONAL_HEADERS
    )
    if unsupported:
        return "应考名单包含不支持的字段，请使用 email、candidate_name 模板"
    return None


def validate_exam_roster_row(row: dict[str, Any]) -> str | None:
    raw_email = _optional_text(row.get("email"))
    if raw_email is not None and len(raw_email) > 255:
        return "邮箱长度不能超过255个字符"
    email = normalize_candidate_email(row.get("email"))
    if email is None:
        return "邮箱不能为空" if raw_email is None else "邮箱格式不正确"
    if len(email) > 255:
        return "邮箱长度不能超过255个字符"
    candidate_name_error = _validate_roster_text(
        row.get("candidate_name"),
        "应考人员姓名",
        max_length=ROSTER_TEXT_MAX_LENGTH,
        required=True,
    )
    if candidate_name_error:
        return candidate_name_error
    for field, label in (
        ("department", "部门"),
        ("position", "职位"),
        ("exam_group", "考试分组"),
    ):
        error = _validate_roster_text(
            row.get(field), label, max_length=ROSTER_TEXT_MAX_LENGTH
        )
        if error:
            return error
    return _validate_roster_text(
        row.get("remark"), "备注", max_length=ROSTER_REMARK_MAX_LENGTH
    )


def _validate_roster_text(
    value: Any, label: str, *, max_length: int, required: bool = False
) -> str | None:
    if value is None:
        return f"{label}不能为空" if required else None
    if not isinstance(value, str):
        return f"{label}格式不正确"
    normalized = value.strip()
    if not normalized:
        return f"{label}不能为空" if required else None
    if len(normalized) > max_length:
        return f"{label}长度不能超过{max_length}个字符"
    if any(ord(char) < 32 for char in normalized):
        return f"{label}包含不支持的控制字符"
    return None


def add_exam_roster_row(
    db: Session, exam_id: int, row: dict[str, Any]
) -> ExamCandidateScope:
    """Create one draft scope, reusing/creating an email-keyed account."""

    exam = db.query(Exam).filter(Exam.id == exam_id).with_for_update().one_or_none()
    if exam is None:
        raise ExamNotFoundError(exam_id)
    if exam.status != "draft":
        raise ExamFrozenError("考试发布后应考名单已冻结")
    reason = validate_exam_roster_row(row)
    if reason:
        raise DomainError(reason)
    email = normalize_candidate_email(row.get("email"))
    assert email is not None
    candidate = _get_or_create_pending_candidate(db, email)
    if candidate.status == "inactive":
        raise DomainError("账号已停用，请先恢复账号")
    existing = (
        db.query(ExamCandidateScope)
        .filter(
            ExamCandidateScope.exam_id == exam_id,
            (ExamCandidateScope.roster_email == email)
            | (ExamCandidateScope.candidate_id == candidate.id),
        )
        .one_or_none()
    )
    if existing is not None:
        raise DomainError("邮箱已在考试名单中")
    try:
        with db.begin_nested():
            scope = ExamCandidateScope(
                exam_id=exam_id,
                candidate_id=candidate.id,
                roster_email=email,
                roster_name=_optional_text(row.get("candidate_name")) or "",
                department=_optional_text(row.get("department")),
                position=_optional_text(row.get("position")),
                exam_group=_optional_text(row.get("exam_group")),
                roster_remark=_optional_text(row.get("remark")),
            )
            db.add(scope)
            db.flush()
    except IntegrityError as exc:
        raise DomainError("邮箱已在考试名单中") from exc
    return scope


def update_exam_roster_row(
    db: Session, scope: ExamCandidateScope, updates: dict[str, Any]
) -> ExamCandidateScope:
    """Apply draft-only scope fields; never mutate the account profile."""

    exam = (
        db.query(Exam).filter(Exam.id == scope.exam_id).with_for_update().one_or_none()
    )
    if exam is None:
        raise ExamNotFoundError(scope.exam_id)
    if exam.status != "draft":
        raise ExamFrozenError("考试发布后应考名单已冻结")
    next_email = updates.get("email", scope.roster_email)
    merged_row = {
        "email": next_email,
        "candidate_name": updates.get("candidate_name", scope.roster_name),
        "department": updates.get("department", scope.department),
        "position": updates.get("position", scope.position),
        "exam_group": updates.get("exam_group", scope.exam_group),
        "remark": updates.get("remark", scope.roster_remark),
    }
    validation_error = validate_exam_roster_row(merged_row)
    if validation_error:
        raise DomainError(validation_error)
    normalized_email = normalize_candidate_email(next_email)
    if normalized_email is None:
        raise DomainError("邮箱格式不正确")
    roster_name = _optional_text(merged_row["candidate_name"])
    assert roster_name is not None
    existing = (
        db.query(ExamCandidateScope)
        .filter(
            ExamCandidateScope.exam_id == scope.exam_id,
            ExamCandidateScope.roster_email == normalized_email,
            ExamCandidateScope.id != scope.id,
        )
        .one_or_none()
    )
    if existing is not None:
        raise DomainError("邮箱已在考试名单中")
    candidate = _get_or_create_pending_candidate(db, normalized_email)
    if candidate.status == "inactive":
        raise DomainError("账号已停用，请先恢复账号")

    scope.roster_email = normalized_email
    scope.roster_name = roster_name
    for field in ("department", "position", "exam_group"):
        if field in updates:
            setattr(scope, field, _optional_text(updates[field]))
    if "remark" in updates:
        scope.roster_remark = _optional_text(updates["remark"])
    # An email update may move a scope to its matching account.  The old
    # account remains intact and is never deleted by a draft edit.
    scope.candidate_id = candidate.id
    db.flush()
    return scope


def _build_pending_candidate(email: str) -> Candidate:
    # Pending accounts intentionally have no display name and cannot receive a
    # candidate token.  The migration makes ``name`` nullable for this state.
    return Candidate(name=None, email=email, status="pending")  # type: ignore[arg-type]


def _get_or_create_pending_candidate(db: Session, email: str) -> Candidate:
    """Race-safe email-keyed account reuse/create using a savepoint."""

    candidate = (
        db.query(Candidate)
        .filter(Candidate.email == email)
        .with_for_update()
        .one_or_none()
    )
    if candidate is not None:
        return candidate
    # A competing transaction may win the normalized-email unique index after
    # the first read.  Retry once when that winner rolled back before the
    # conflict became visible; otherwise lock and reuse the winner.
    for _attempt in range(2):
        try:
            with db.begin_nested():
                candidate = _build_pending_candidate(email)
                db.add(candidate)
                db.flush()
            return candidate
        except IntegrityError:
            candidate = (
                db.query(Candidate)
                .filter(Candidate.email == email)
                .with_for_update()
                .one_or_none()
            )
            if candidate is not None:
                # The winner is authoritative; never merge or overwrite its
                # display name or lifecycle status.
                return candidate
    raise DomainError("邮箱账号创建冲突，请重试")


def import_exam_roster_from_workbook(
    db: Session,
    exam_id: int,
    file_obj: Any,
    file_name: str,
    *,
    commit: bool = True,
) -> QuestionImportResult:
    """Import a bounded email-keyed roster into one draft exam.

    File size, sheet count, row count, and headers are validated before any
    account/scope mutation.  Each row then either succeeds atomically or is
    represented in the batch failure report; invalid/inactive identities never
    receive a partial scope.
    """

    assert_admin_mutation_allowed(db)
    exam = db.query(Exam).filter(Exam.id == exam_id).with_for_update().one_or_none()
    if exam is None:
        raise ExamNotFoundError(exam_id)
    if exam.status != "draft":
        raise ExamFrozenError("考试发布后应考名单已冻结")

    validate_upload_file_size(file_obj)
    parsed = parse_workbook(file_obj)
    header_error = validate_exam_roster_headers(parsed.headers)
    if header_error:
        raise DomainError(header_error)

    failures: list[ImportFailure] = []
    success_count = 0
    seen_emails: set[str] = set()

    for row_number, row in enumerate(parsed.rows, start=2):
        reason = validate_exam_roster_row(row)
        if reason:
            failures.append(ImportFailure(row_number=row_number, reason=reason))
            continue
        email = normalize_candidate_email(row.get("email"))
        assert email is not None  # validate_exam_roster_row established this
        if email in seen_emails:
            failures.append(
                ImportFailure(row_number=row_number, reason="邮箱在本批次重复")
            )
            continue
        seen_emails.add(email)

        try:
            candidate = _get_or_create_pending_candidate(db, email)
        except DomainError as exc:
            failures.append(ImportFailure(row_number=row_number, reason=str(exc)))
            continue
        if candidate.status == "inactive":
            failures.append(
                ImportFailure(row_number=row_number, reason="账号已停用，请先恢复账号")
            )
            continue

        existing_scope = (
            db.query(ExamCandidateScope)
            .filter(
                ExamCandidateScope.exam_id == exam_id,
                (
                    (ExamCandidateScope.roster_email == email)
                    | (ExamCandidateScope.candidate_id == candidate.id)
                ),
            )
            .one_or_none()
        )
        if existing_scope is not None:
            failures.append(
                ImportFailure(row_number=row_number, reason="邮箱已在考试名单中")
            )
            continue

        try:
            # The exam row lock serializes same-exam imports, while this
            # savepoint converts a concurrent unique-scope race into a stable
            # row-level failure instead of aborting the whole workbook.
            with db.begin_nested():
                scope = ExamCandidateScope(
                    exam_id=exam_id,
                    candidate_id=candidate.id,
                    roster_email=email,
                    roster_name=_optional_text(row.get("candidate_name")) or "",
                    department=_optional_text(row.get("department")),
                    position=_optional_text(row.get("position")),
                    exam_group=_optional_text(row.get("exam_group")),
                    roster_remark=_optional_text(row.get("remark")),
                )
                db.add(scope)
                db.flush()
        except IntegrityError:
            failures.append(
                ImportFailure(row_number=row_number, reason="邮箱已在考试名单中")
            )
            continue
        success_count += 1

    batch = ImportBatch(
        import_type="exam_candidates",
        file_name=file_name,
        total_count=parsed.total_count,
        success_count=success_count,
        failed_count=len(failures),
        status="completed",
        error_report=[failure.model_dump() for failure in failures],
    )
    db.add(batch)
    db.flush()
    if commit:
        db.commit()
    return QuestionImportResult(
        batch_id=batch.id,
        success_count=success_count,
        failed_count=len(failures),
        failures=failures,
    )


def generate_failure_report(db: Session, batch_id: int) -> BytesIO:
    batch = db.get(ImportBatch, batch_id)
    if batch is None:
        raise ImportBatchNotFoundError(batch_id)

    workbook = Workbook()
    meta_sheet = workbook.active
    meta_sheet.title = "导入批次"
    meta_sheet.append(["字段", "值"])
    meta_sheet.append(
        ["导入类型", IMPORT_TYPE_LABELS.get(batch.import_type, batch.import_type)]
    )
    meta_sheet.append(["文件名", escape_excel_cell(batch.file_name)])
    meta_sheet.append(["总数", batch.total_count])
    meta_sheet.append(["成功数", batch.success_count])
    meta_sheet.append(["失败数", batch.failed_count])
    meta_sheet.append(["生成时间", datetime.now(UTC).isoformat()])

    detail_sheet = workbook.create_sheet("失败明细")
    detail_sheet.append(["ROW · 行号", "REASON · 原因"])
    for failure in batch.error_report or []:
        detail_sheet.append(
            [
                escape_excel_cell(failure.get("row_number")),
                escape_excel_cell(failure.get("reason")),
            ]
        )

    for sheet in workbook.worksheets:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(
                max(max_length + 2, 10), 48
            )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def validate_question_import_row(row: dict[str, Any]) -> str | None:
    question_type = _text(row.get("question_type")).lower()
    stem = _text(row.get("stem"))
    status = _text(row.get("status") or DEFAULT_STATUS).lower()
    correct_answer = _text(row.get("correct_answer"))

    if not question_type:
        return "题型不能为空"
    if question_type not in VALID_QUESTION_TYPES:
        return "题型只能填写单选（single）、多选（multiple）或判断（judge）"
    if not stem:
        return "题干不能为空"
    if status not in VALID_STATUSES:
        return "状态只能填写启用（active）或停用（inactive）"
    if not _is_number(row.get("score")):
        return "分值必须是数字"

    answers = _parse_correct_option_labels(question_type, correct_answer)
    if question_type == "single" and len(answers) != 1:
        return "单选题只能有一个正确答案"
    if question_type == "multiple" and len(answers) < 2:
        return "多选题至少两个正确答案"
    if question_type == "judge" and correct_answer.lower() not in {"true", "false"}:
        return "判断题答案只能是 true 或 false"

    if question_type in {"single", "multiple"}:
        existing_labels = {
            label
            for label in OPTION_LABELS
            if _optional_text(row.get(f"option_{label.lower()}"))
        }
        missing = [answer for answer in answers if answer not in existing_labels]
        if missing:
            return "正确答案必须存在于选项中"
    return None


def _build_question(row: dict[str, Any]) -> Question:
    question_type = _text(row.get("question_type")).lower()
    correct_answers = _parse_correct_option_labels(
        question_type, _text(row.get("correct_answer"))
    )
    question = Question(
        question_type=question_type,
        stem=_text(row.get("stem")),
        analysis=_optional_text(row.get("analysis")),
        category_1=_optional_text(row.get("category_1")),
        category_2=_optional_text(row.get("category_2")),
        difficulty=_optional_text(row.get("difficulty")),
        score=Decimal(_text(row.get("score"))),
        status=_text(row.get("status") or DEFAULT_STATUS).lower(),
        source=_optional_text(row.get("source")),
        source_no=_optional_text(row.get("source_no")),
        remark=_optional_text(row.get("remark")),
    )
    question.options = [
        QuestionOption(
            label=label,
            content=content,
            is_correct=label in correct_answers,
            sort_order=index,
        )
        for index, (label, content) in enumerate(
            _extract_options(row, question_type), start=1
        )
    ]
    return question


def normalize_candidate_email(raw_email: object) -> str | None:
    try:
        email = normalize_email(raw_email)
    except ValueError:
        return None
    try:
        return str(EMAIL_ADAPTER.validate_python(email)).lower()
    except ValidationError:
        return None


def _parse_correct_option_labels(question_type: str, raw: str) -> set[str]:
    if question_type == "judge":
        label = JUDGE_ANSWER_MAP.get(raw.strip().lower())
        if label is not None:
            return {label}
    return normalize_answer_set(raw)


def _extract_options(row: dict[str, Any], question_type: str) -> list[tuple[str, str]]:
    options = [
        (label, content)
        for label in OPTION_LABELS
        if (content := _optional_text(row.get(f"option_{label.lower()}"))) is not None
    ]
    if question_type == "judge" and not options:
        return list(JUDGE_OPTIONS)
    return options


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _optional_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _is_number(value: Any) -> bool:
    try:
        float(value)
    except (TypeError, ValueError):
        return False
    return True
